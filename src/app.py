from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import pyodbc
from config import config
from models.modelUser import ModelUser
from models.entities.User import User
import matplotlib.pyplot as plt
import io
import base64
import os
from utils.chart_generator import generar_grafico
from flask_paginate import Pagination, get_page_args
from flask import Flask, render_template, request
import openpyxl


app = Flask(__name__)
app.config.from_object(config['development'])


connection_string = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={app.config['MSSQL_SERVER']};"
    f"DATABASE={app.config['MSSQL_DATABASE']};"
    f"UID={app.config['MSSQL_USERNAME']};"
    f"PWD={app.config['MSSQL_PASSWORD']};"
    f"Trusted_Connection={app.config['MSSQL_TRUSTED_CONNECTION']}"
)
db = pyodbc.connect(connection_string)
print("Conexión exitosa a la base de datos")


@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User(0, username, password)
        logged_user = ModelUser.login(db, user)
        if logged_user is not None:
            if logged_user.password == user.password:
                return redirect(url_for('home'))
            else:
                flash("Contraseña inválida")
                return render_template('auth/login.html')
        else:
            flash("Usuario no encontrado")
            return render_template('auth/login.html')
    else:
        return render_template('auth/login.html')


@app.route('/home')
def home():
    return render_template('auth/home.html')


@app.route('/customer')
def customer():
    # Obtener los datos del procedimiento almacenado ObtenerInformacionClientes
    cursor = db.cursor()
    # ----------CAMBIAR A PROCEDIMEINTO ALMACENADO
    cursor.execute("SELECT * FROM Cliente")
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]

    # Configurar la paginación
    page = request.args.get('page', 1, type=int)
    per_page = 8  # Número de registros por página
    total = len(rows)

    pagination = Pagination(page=page, per_page=per_page,
                            total=total, css_framework='bootstrap4')

    # Calcular los índices de inicio y fin de los registros a mostrar
    start = (page - 1) * per_page
    end = min(start + per_page, total)

    return render_template('auth/customer.html', columns=columns, rows=rows[start:end], pagination=pagination)


@app.route('/products')
def products():
    # Obtener los datos del procedimiento almacenado ObtenerInformacionProductos
    cursor = db.cursor()
    # ----------CAMBIAR A PROCEDIMEINTO ALMACENADO
    cursor.execute("SELECT * FROM Producto")
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]

    # Configurar la paginación
    page = request.args.get('page', 1, type=int)
    per_page = 8  # Número de registros por página
    total = len(rows)

    pagination = Pagination(page=page, per_page=per_page,
                            total=total, css_framework='bootstrap4')

    # Calcular los índices de inicio y fin de los registros a mostrar
    start = (page - 1) * per_page
    end = min(start + per_page, total)

    return render_template('auth/products.html', columns=columns, rows=rows[start:end], pagination=pagination)


@app.route('/register')
def register():
    return render_template('auth/register.html')


@app.route('/divisionales')
def divisionales():

    # Obtener los datos del procedimiento almacenado ObtenerInformacionProductos
    cursor = db.cursor()
    # ---------- CAMBIAR A PROCEDIMIENTO ALMACENADO
    cursor.execute("SELECT * FROM Divisional")
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]

    # Configurar la paginación
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Número de registros por página
    total = len(rows)

    pagination = Pagination(page=page, per_page=per_page,
                            total=total, css_framework='bootstrap4')

    # Calcular los índices de inicio y fin de los registros a mostrar
    start = (page - 1) * per_page
    end = min(start + per_page, total)

    return render_template('auth/divisionales.html', columns=columns, rows=rows[start:end], pagination=pagination)


@app.route('/logout')
def logout():
    return render_template('auth/login.html')


@app.route('/supervisores')
def supervisores():

    # Obtener los datos del procedimiento almacenado ObtenerInformacionProductos
    cursor = db.cursor()
    # --------CAMBIAR A PROCEDIMEINTO ALMACENADO
    cursor.execute("SELECT * FROM Supervisor")
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]

    # Configurar la paginación
    page = request.args.get('page', 1, type=int)
    per_page = 8  # Número de registros por página
    total = len(rows)

    pagination = Pagination(page=page, per_page=per_page,
                            total=total, css_framework='bootstrap4')

    # Calcular los índices de inicio y fin de los registros a mostrar
    start = (page - 1) * per_page
    end = min(start + per_page, total)

    return render_template('auth/supervisores.html', columns=columns, rows=rows[start:end], pagination=pagination)


def obtener_nombre(id):
    # Lógica para obtener el nombre correspondiente al ID
    # Ejemplo: Obtener el nombre de una base de datos o cualquier otra fuente de datos
    # Puedes adaptar esta función según tus necesidades específicas
    if id == '1':
        return 'Nombre 1'
    elif id == '2':
        return 'Nombre 2'
    elif id == '3':
        return 'Nombre 3'
    else:
        return 'Nombre desconocido'


@app.route('/editar_customer')
def editar_customer():
    id = request.args.get('id')  # Obtener el ID de la fila desde la URL
    # Obtener los datos de la fila correspondiente utilizando el ID
    # Ejemplo: Obtener el nombre utilizando la función obtener_nombre
    nombre = obtener_nombre(id)
    return render_template('auth/edit_customer.html', id=id, nombre=nombre)


@app.route('/editar_products')
def editar_products():
    id = request.args.get('id')  # Obtener el ID de la fila desde la URL
    # Obtener los datos de la fila correspondiente utilizando el ID
    # Ejemplo: Obtener el nombre utilizando la función obtener_nombre
    nombre = obtener_nombre(id)
    return render_template('auth/edit_products.html', id=id, nombre=nombre)


@app.route('/editar_supervisores')
def editar_supervisores():
    id = request.args.get('id')  # Obtener el ID de la fila desde la URL
    # Obtener los datos de la fila correspondiente utilizando el ID
    # Ejemplo: Obtener el nombre utilizando la función obtener_nombre
    nombre = obtener_nombre(id)
    return render_template('auth/edit_supervisores.html', id=id, nombre=nombre)


@app.route('/editar_divisionales')
def editar_divisionales():
    id = request.args.get('id')  # Obtener el ID de la fila desde la URL
    # Obtener los datos de la fila correspondiente utilizando el ID
    # Ejemplo: Obtener el nombre utilizando la función obtener_nombre
    nombre = obtener_nombre(id)
    return render_template('auth/edit_divisionales.html', id=id, nombre=nombre)


@app.route('/descargar_base_customer', methods=['GET'])
def descargar_base_customer():
    # Establecer la conexión a la base de datos
    cursor = db.cursor()
    # Ejecutar el procedimiento almacenado para obtener los datos
    cursor.execute("SELECT * FROM Cliente")
    results = cursor.fetchall()

    # Crear un nuevo archivo XLSX utilizando openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar los encabezados de columna
    headers = [column[0] for column in cursor.description]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Agregar los datos obtenidos del procedimiento almacenado a la hoja de cálculo
    for row_num, row in enumerate(results, 2):
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Guardar el archivo XLSX en un directorio temporal
    # Reemplaza esto con la ruta y nombre de archivo deseados
    filename = 'C:/Users/jonat/OneDrive/Escritorio/App Budget/downloads/clientesAass.xlsx'
    workbook.save(filename)

    # Devolver el archivo XLSX como una descarga al usuario
    return send_file(filename, as_attachment=True)


@app.route('/descargar_base_products', methods=['GET'])
def descargar_base_products():
    # Establecer la conexión a la base de datos
    cursor = db.cursor()
    # Ejecutar el procedimiento almacenado para obtener los datos
    cursor.execute("SELECT * FROM Producto")
    results = cursor.fetchall()

    # Crear un nuevo archivo XLSX utilizando openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar los encabezados de columna
    headers = [column[0] for column in cursor.description]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Agregar los datos obtenidos del procedimiento almacenado a la hoja de cálculo
    for row_num, row in enumerate(results, 2):
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Guardar el archivo XLSX en un directorio temporal
    # Reemplaza esto con la ruta y nombre de archivo deseados
    filename = 'C:/Users/jonat/OneDrive/Escritorio/App Budget/downloads/ProductosIdeal.xlsx'
    workbook.save(filename)

    # Devolver el archivo XLSX como una descarga al usuario
    return send_file(filename, as_attachment=True)


@app.route('/descargar_base_supervisores', methods=['GET'])
def descargar_base_supervisores():
    # Establecer la conexión a la base de datos
    cursor = db.cursor()
    # Ejecutar el procedimiento almacenado para obtener los datos
    cursor.execute("SELECT * FROM Supervisor")
    results = cursor.fetchall()

    # Crear un nuevo archivo XLSX utilizando openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar los encabezados de columna
    headers = [column[0] for column in cursor.description]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Agregar los datos obtenidos del procedimiento almacenado a la hoja de cálculo
    for row_num, row in enumerate(results, 2):
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Guardar el archivo XLSX en un directorio temporal
    # Reemplaza esto con la ruta y nombre de archivo deseados
    filename = 'C:/Users/jonat/OneDrive/Escritorio/App Budget/downloads/SupervisoresAass.xlsx'
    workbook.save(filename)

    # Devolver el archivo XLSX como una descarga al usuario
    return send_file(filename, as_attachment=True)


@app.route('/descargar_base_divisionales', methods=['GET'])
def descargar_base_divisionales():
    # Establecer la conexión a la base de datos
    cursor = db.cursor()
    # Ejecutar el procedimiento almacenado para obtener los datos
    cursor.execute("SELECT * FROM Divisional")
    results = cursor.fetchall()

    # Crear un nuevo archivo XLSX utilizando openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar los encabezados de columna
    headers = [column[0] for column in cursor.description]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Agregar los datos obtenidos del procedimiento almacenado a la hoja de cálculo
    for row_num, row in enumerate(results, 2):
        for col_num, value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Guardar el archivo XLSX en un directorio temporal
    # Reemplaza esto con la ruta y nombre de archivo deseados
    filename = 'C:/Users/jonat/OneDrive/Escritorio/App Budget/downloads/DivisionalesAass.xlsx'
    workbook.save(filename)

    # Devolver el archivo XLSX como una descarga al usuario
    return send_file(filename, as_attachment=True)


if __name__ == '__main__':
    app.run()
